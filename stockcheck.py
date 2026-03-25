from __future__ import annotations

import csv
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import pandas as pd
from tomlkit import ws
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from datetime import datetime, timedelta
from pandas.tseries.holiday import USFederalHolidayCalendar
from pandas.tseries.offsets import CustomBusinessDay


INPUT_FILE = "stocks.csv"
OUTPUT_FILE = "results.xlsx"

RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")     #Current Price < Target Low
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFACD")  #Current Price > Target High

def get_last_business_date():
    us_bd = CustomBusinessDay(calendar=USFederalHolidayCalendar())
    last_business_day = pd.Timestamp.today().normalize() - us_bd
    return last_business_day.strftime('%Y-%m-%d')

def fetch_closed_price(symbol):
    last_business_date = get_last_business_date()
    today = (datetime.now() + timedelta(days=0)).strftime('%Y-%m-%d')
    data = yf.download(symbol, start=last_business_date, end=today, progress=False)

    if not data.empty:
        closing_price = data['Close'].iloc[0].item()  # Ensure it's a scalar float
        #print(f"Closing price for {symbol} on {last_business_date}: ${closing_price:.2f}")
        return closing_price
    else:
        print("No data available for the specified date.")
        return None

def parse_money(value: str) -> Optional[float]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("$", "").replace(",", "")

    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None


def fetch_prices_batch(symbols: list[str], max_retries: int = 3, delay: int = 2) -> dict[str, Optional[float]]:
    """
    Fetch latest close prices for all symbols in one yfinance call.
    Returns a dict like {"AAPL": 213.49, "SPY": 589.22, "USO": 81.34}
    If a symbol cannot be fetched, it will remain None.
    """
    cleaned = []
    seen = set()

    for s in symbols:
        sym = str(s).strip().upper()
        if sym and sym not in seen:
            cleaned.append(sym)
            seen.add(sym)

    prices = {sym: None for sym in cleaned}
    if not cleaned:
        return prices

    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            df = yf.download(
                tickers=cleaned,
                period="5d",
                interval="1d",
                group_by="ticker",
                auto_adjust=False,
                progress=False,
                threads=False,
            )

            if df is None or df.empty:
                raise RuntimeError("No data returned from yfinance")

            # Case 1: multiple tickers -> MultiIndex columns: (Ticker, Field)
            if isinstance(df.columns, pd.MultiIndex):
                for sym in cleaned:
                    try:
                        if sym in df.columns.get_level_values(0):
                            close_series = df[sym]["Close"].dropna()
                            if not close_series.empty:
                                prices[sym] = float(close_series.iloc[-1])
                    except Exception:
                        pass

            # Case 2: single ticker -> normal columns: Open, High, Low, Close...
            else:
                if len(cleaned) == 1 and "Close" in df.columns:
                    close_series = df["Close"].dropna()
                    if not close_series.empty:
                        prices[cleaned[0]] = float(close_series.iloc[-1])

            return prices

        except Exception as e:
            last_error = e
            if attempt < max_retries:
                sleep_seconds = delay 
                print(f"Batch fetch attempt {attempt} failed: {e}")
                print(f"Retrying in {sleep_seconds} seconds...")
                time.sleep(sleep_seconds)

    print(f"Batch fetch failed after {max_retries} attempts: {last_error}")
    return prices


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def main() -> None:
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    with input_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        raise ValueError("The input CSV is empty.")

    header = rows[0]
    if len(header) < 3:
        raise ValueError(
            "The input CSV must have at least 3 columns: "
            "'Stocks/ETF', 'Target Low', 'Target High'."
        )

    # Gather all symbols once
    symbols = []
    for row in rows[1:]:
        if row and len(row) > 0:
            symbols.append(row[0])

    price_map = fetch_prices_batch(symbols)

    output_header = header[:3] + [
    "Current Price",
    "Previous Closing Price",
    "Today's Change $",
    "Today's Change %"
    ]  + header[3:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Check"
    ws.append(output_header)

    bold_font = Font(bold=True)

    for col in range(1, len(output_header) + 1):
        ws.cell(row=1, column=col).font = bold_font


    ws.freeze_panes = "A2"
    for col in range(1, len(output_header) + 1):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    output_data = []

    for row in rows[1:]:
        padded_row = row + [""] * max(0, len(header) - len(row))

        symbol = str(padded_row[0]).strip().upper()
        target_low = parse_money(padded_row[1])
        target_high = parse_money(padded_row[2])
        current_price = price_map.get(symbol)
        previous_closing_price = fetch_closed_price(symbol) if symbol else None
        change_dollar = None
        change_percent = None

        if (
            isinstance(current_price, (int, float))
            and isinstance(previous_closing_price, (int, float))
            and previous_closing_price != 0
        ):
            change_dollar = current_price - previous_closing_price
            change_percent = (change_dollar / previous_closing_price) * 100

        current_price_output = current_price if current_price is not None else "NO DATA"
        previous_closing_price_output = (
            previous_closing_price if previous_closing_price is not None else "NO DATA"
        )
        change_dollar_output = change_dollar if change_dollar is not None else "NO DATA"
        change_percent_output = (
            change_percent / 100 if change_percent is not None else "NO DATA"
        )


        output_row = (
            padded_row[:3]
            + [
            current_price_output,
            previous_closing_price_output,
            change_dollar_output,
            change_percent_output,
            ]
            + padded_row[3:]
        )



        output_data.append({
            "row": output_row,
            "change_percent": change_percent if isinstance(change_percent, (int, float)) else float("-inf"),
            "target_low": target_low,
            "target_high": target_high,
        })

        excel_row = ws.max_row
 
    
    output_data.sort(key=lambda x: x["change_percent"], reverse=True)

    for item in output_data:
        ws.append(item["row"])
        excel_row = ws.max_row

        current_price = item["row"][3]
        previous_closing_price = item["row"][4]
        change_dollar = item["row"][5]
        change_percent = item["row"][6]

        target_low = item["target_low"]
        target_high = item["target_high"]

        ws.cell(excel_row, 2).number_format = "$#,##0.00"
        ws.cell(excel_row, 3).number_format = "$#,##0.00"

        if isinstance(current_price, (int, float)):
            ws.cell(excel_row, 4).number_format = "$#,##0.00"

        if isinstance(previous_closing_price, (int, float)):
            ws.cell(excel_row, 5).number_format = "$#,##0.00"

        if isinstance(change_dollar, (int, float)):
            ws.cell(excel_row, 6).number_format = "$#,##0.00"

        if isinstance(change_percent, (int, float)):
            ws.cell(excel_row, 7).number_format = "0.00%"

        fill = None

        if (
            isinstance(current_price, (int, float))
            and target_low is not None
            and target_high is not None
        ):
            if current_price < target_low:
                fill = RED_FILL
            elif current_price > target_high:
                fill = YELLOW_FILL

        if fill:
            for col in range(1, len(output_header) + 1):
                ws.cell(excel_row, col).fill = fill

    autosize_columns(ws)

    # add borders here
    from openpyxl.styles import Border, Side

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


    wb.save(OUTPUT_FILE)
    print(f"Done. Output written to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
